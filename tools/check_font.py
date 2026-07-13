#!/usr/bin/env python3
"""
Tool kiểm tra unicode tổ hợp (dấu rời/NFD) trong text hoặc file Excel.

"Unicode tổ hợp" = ký tự được tạo từ chữ cái gốc + dấu kết hợp riêng biệt
(dạng NFD), khác với "unicode dựng sẵn" (dạng NFC) - nguyên nhân phổ biến
gây lỗi hiển thị font tiếng Việt.
"""

import re
import unicodedata

MDV2_SPECIAL_RE = re.compile(r"([_*\[\]()~`>#+\-=|{}.!\\])")

# Ký tự "nhìn giống nhưng sai mã" thường gặp khi convert font tiếng Việt cũ
# (TCVN3/VNI) sang Unicode: ví dụ Đ (U+0110) bị mã nhầm thành Ð - chữ Eth
# (U+00D0), đ (U+0111) bị mã nhầm thành ð (U+00F0). Về mặt Unicode đây là 2
# ký tự hoàn toàn khác nhau (không phải NFC/NFD) nên phải bắt riêng.
CONFUSABLE_CHARS = {
    "Ð": "Đ",  # Ð -> Đ
    "ð": "đ",  # ð -> đ
}


def has_to_hop(text: str) -> bool:
    """True nếu text chứa unicode tổ hợp (khác dạng chuẩn hóa NFC)."""
    if not isinstance(text, str) or not text:
        return False
    return unicodedata.normalize("NFC", text) != text


def has_confusable_char(text: str) -> bool:
    """True nếu text chứa ký tự bị mã nhầm (VD: Ð/ð thay cho Đ/đ)."""
    if not isinstance(text, str) or not text:
        return False
    return any(ch in CONFUSABLE_CHARS for ch in text)


def has_font_error(text: str) -> bool:
    """True nếu text có lỗi font tiếng Việt: unicode tổ hợp hoặc ký tự bị mã nhầm."""
    return has_to_hop(text) or has_confusable_char(text)


def escape_mdv2(text: str) -> str:
    """Escape ký tự đặc biệt cho Telegram MarkdownV2."""
    return MDV2_SPECIAL_RE.sub(r"\\\1", text)


def iter_grapheme_clusters(text: str):
    """
    Tách text thành các cụm ký tự: 1 chữ cái gốc + các dấu kết hợp (combining
    mark) đi theo ngay sau nó. Mỗi cụm là 1 "ký tự nhìn thấy" thực tế.
    """
    n = len(text)
    i = 0
    while i < n:
        j = i + 1
        while j < n and unicodedata.combining(text[j]) != 0:
            j += 1
        yield text[i:j]
        i = j


def underline_to_hop_chars(text: str) -> str:
    """
    Trả về chuỗi MarkdownV2 (đã escape) với các ký tự unicode tổ hợp được
    gạch dưới (__..__), giữ nguyên các ký tự bình thường.
    """
    if not text:
        return text
    out = []
    for cluster in iter_grapheme_clusters(text):
        escaped = escape_mdv2(cluster)
        is_error = any(unicodedata.combining(ch) != 0 for ch in cluster) or any(
            ch in CONFUSABLE_CHARS for ch in cluster
        )
        out.append(f"__{escaped}__" if is_error else escaped)
    return "".join(out)


def check_excel_file(input_path: str, output_path: str) -> dict:
    """
    Quét toàn bộ ô trong file Excel, tô nền đỏ các ô chứa unicode tổ hợp,
    lưu kết quả ra output_path.

    Trả về: {"total_cells": int, "flagged_cells": int, "sheets": {sheet_name: count}}
    """
    import openpyxl
    from openpyxl.styles import PatternFill

    wb = openpyxl.load_workbook(input_path)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

    total_cells = 0
    flagged_cells = 0
    flagged_by_sheet = {}

    for ws in wb.worksheets:
        sheet_flagged = 0
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                total_cells += 1
                if isinstance(value, str) and has_font_error(value):
                    cell.fill = red_fill
                    flagged_cells += 1
                    sheet_flagged += 1
        if sheet_flagged:
            flagged_by_sheet[ws.title] = sheet_flagged

    wb.save(output_path)
    return {
        "total_cells": total_cells,
        "flagged_cells": flagged_cells,
        "sheets": flagged_by_sheet,
    }
