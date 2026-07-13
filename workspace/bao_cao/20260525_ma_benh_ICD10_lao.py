import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

data = [
    (1, "A15", "A15: Lao hô hấp, có xác nhận về vi khuẩn học và mô học"),
    (2, "A16", "A16: Lao đường hô hấp, không xác nhận về vi khuẩn học hoặc mô học"),
    (3, "A17", "A17: Lao hệ thần kinh"),
    (4, "A18", "A18: Lao các cơ quan khác"),
    (5, "A19", "A19: Lao kê"),
    (6, "B90", "B90: Di chứng do lao"),
    (7, "J65", "J65: Bệnh bụi phổi kết hợp với lao"),
    (8, "K23.0", "K23.0*: Viêm thực quản do lao (A18.8†)"),
    (9, "K67.3", "K67.3*: Viêm phúc mạc do lao (A18.3†)"),
    (10, "K93.0", "K93.0*: Bệnh lao ở ruột, phúc mạc và tuyến mạc treo (A18.3†)"),
    (11, "M01.1*", "M01.1*: Viêm khớp do lao (A18.0†)"),
    (12, "M49.0*", "M49.0*: Lao cột sống (A18.0†)"),
    (13, "M90.0*", "M90.0*: Lao xương (A18.0†)"),
    (14, "N33.0", "N33.0: Lao bàng quang (A18.1†)"),
    (15, "N74.0*", "N74.0*: Lao cổ tử cung (A18.1†)"),
    (16, "N74.1*", "N74.1*: Viêm lao vùng chậu nữ (A18.1†)"),
    (17, "O98.0", "O98.0: Bệnh lao gây biến chứng cho thai kỳ, khi đẻ và sau đẻ"),
    (18, "P37.0", "P37.0: Lao bẩm sinh"),
    (19, "U84.3", "U84.3: Kháng (các) thuốc chống lao"),
    (20, "Z03.0", "Z03.0: Theo dõi, đánh giá người tự nghi ngờ mắc bệnh lao"),
    (21, "Z11.1", "Z11.1: Khám sàng lọc chuyên khoa về lao phổi"),
    (22, "Z20.1", "Z20.1: Tiếp xúc và phơi nhiễm với vi khuẩn lao"),
    (23, "Z23.2", "Z23.2: Tiêm phòng Lao (BCG)"),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mã bệnh ICD10"

# Styles
header_font = Font(name="Times New Roman", bold=True, size=12)
cell_font = Font(name="Times New Roman", size=11)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Header
ws.merge_cells("A1:C1")
title_cell = ws["A1"]
title_cell.value = "DANH MỤC MÃ BỆNH LAO THEO ICD 10"
title_cell.font = Font(name="Times New Roman", bold=True, size=13)
title_cell.alignment = center_align
ws.row_dimensions[1].height = 30

# Column headers
headers = ["TT", "Mã bệnh theo ICD 10", "Bệnh/tình trạng"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = border
ws.row_dimensions[2].height = 30

# Data rows
for row_idx, (tt, ma, benh) in enumerate(data, 3):
    ws.cell(row=row_idx, column=1, value=tt).alignment = center_align
    ws.cell(row=row_idx, column=2, value=ma).alignment = center_align
    ws.cell(row=row_idx, column=3, value=benh).alignment = left_align
    for col in range(1, 4):
        ws.cell(row=row_idx, column=col).font = cell_font
        ws.cell(row=row_idx, column=col).border = border
    ws.row_dimensions[row_idx].height = 25

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 70

output = "/Users/hung/Projects/private/tieu_me/workspace/bao_cao/20260525_ma_benh_ICD10_lao.xlsx"
wb.save(output)
print(f"Saved: {output}")
